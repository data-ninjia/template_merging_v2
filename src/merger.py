from copy import copy
from openpyxl import Workbook


class ExcelMerger:
    def __init__(self, output_name: str):
        self.output_name = output_name
        self.wb = Workbook()
        self.ws = self.wb.active
        self.current_row = 1
        self.header_added = False
        self.target_max_col = 0

    def append_data(self, source_ws, file_name: str):
        if not self.header_added:
            self.target_max_col = source_ws.max_column
            for col_letter, col_dim in source_ws.column_dimensions.items():
                self.ws.column_dimensions[col_letter].width = col_dim.width
                self.ws.column_dimensions[col_letter].hidden = False

        for r_idx, row in enumerate(
            source_ws.iter_rows(max_col=self.target_max_col), 1
        ):
            if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
                continue

            if r_idx == 1:
                if not self.header_added:
                    for cell in row:
                        new_cell = self.ws.cell(
                            row=self.current_row, column=cell.column
                        )
                        self._copy_cell(cell, new_cell)

                    source_header = self.ws.cell(
                        row=self.current_row, column=self.target_max_col + 1
                    )
                    source_header.value = "Source File"
                    self._copy_cell(row[-1], source_header, only_style=True)

                    self.header_added = True
                    self.current_row += 1
                continue

            for cell in row:
                new_cell = self.ws.cell(row=self.current_row, column=cell.column)
                self._copy_cell(cell, new_cell)

            self.ws.cell(row=self.current_row, column=self.target_max_col + 1).value = (
                file_name
            )
            self.current_row += 1

    def _copy_cell(self, src, tgt, only_style=False):
        if not only_style:
            tgt.value = src.value
            if src.value is not None:
                tgt.data_type = src.data_type

        if src.has_style:
            tgt.font = copy(src.font)
            tgt.border = copy(src.border)
            tgt.fill = copy(src.fill)
            tgt.number_format = copy(src.number_format)
            tgt.protection = copy(src.protection)
            tgt.alignment = copy(src.alignment)

    def save(self):
        self.wb.save(self.output_name)
